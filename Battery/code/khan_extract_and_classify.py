"""
Phase 4 confirmation run — Khan 2025 held-out cohort.

Per the pre-registration at literature/09_phase4_pre_registration.md:
  Features:   per-trajectory residual-direction unit vector at flagged RPTs
  Centroids:  LLI = (-1, 0, 0); LAM+SEI = (-1, +1, +1)/sqrt(3)
  Classifier: cosine similarity to centroids, confidence threshold 0.3
  Falsification:
    1. >=50% of cells confidently classified
    2. >=70% of classified cells are LAM+SEI (NMC chemistry expectation)
    3. Permutation null p<0.05 (Bonferroni-corrected /3 for the three replication cohorts)

Khan 2025 cohort:
  Cells: ~22 valid (S2, S18 excluded per Khan paper data-collection errors)
  Operators: Q_max, R_ohmic, R_diff (EIS bands at mid-SOC = S50)
  Timepoints: days 0, 10, 20, 40, 90 (5 RPTs cycle-aged); days 0, 70, 90 (calendar)
  Fresh ref: day 0 (each cell's own day-0 observation; pooled covariance across cells)

This script is the locked classifier run. Report regardless of direction.
"""

from pathlib import Path
import re
import numpy as np
import pandas as pd
import openpyxl
from scipy import stats


DATA = Path("D:/Renewables/Battery/data/khan_2025")
OUT_DIR = Path("D:/Renewables/Battery/data/processed")
EXCLUDE_CELLS = {"S2", "S18"}

# Pre-registered centroids
U_LLI = np.array([-1.0, 0.0, 0.0])
U_LLI = U_LLI / np.linalg.norm(U_LLI)
U_LAM_SEI = np.array([-1.0, +1.0, +1.0])
U_LAM_SEI = U_LAM_SEI / np.linalg.norm(U_LAM_SEI)
CONF_THRESHOLD = 0.3
DAYS_CYCLE = [0, 10, 20, 40, 90]


def extract_qmax_per_cell_per_day():
    """Walk CD_t25_*d.xlsx files, return DataFrame [cell, day, Q_max]."""
    rows = []
    for day in DAYS_CYCLE:
        path = DATA / "capacity" / f"CD_t25_{day}d.xlsx"
        if not path.exists():
            continue
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        for sn in wb.sheetnames:
            if not re.match(r"^S\d+$", sn) or sn in EXCLUDE_CELLS:
                continue
            ws = wb[sn]
            max_cdcg = -np.inf
            for row in ws.iter_rows(min_row=7, values_only=True):
                if len(row) > 6:
                    v = row[6]  # Cdcg column
                    if v is not None and isinstance(v, (int, float)) and v > max_cdcg:
                        max_cdcg = v
            if max_cdcg != -np.inf:
                rows.append({"cell_id": sn, "day": day, "Q_max_Ah": float(max_cdcg)})
    return pd.DataFrame(rows)


def extract_eis_per_cell_per_day(soc="S50"):
    """Walk eis_csv/S*/ACR_t25_{soc}_{day}d_S{n}_convert.csv, extract R_ohmic + R_diff."""
    rows = []
    eis_root = DATA / "eis_csv"
    for cell_dir in eis_root.iterdir():
        if not cell_dir.is_dir():
            continue
        cell = cell_dir.name  # like 'S1'
        if cell in EXCLUDE_CELLS:
            continue
        for day in DAYS_CYCLE:
            fname = f"ACR_t25_{soc}_{day}d_{cell}_convert.csv"
            path = cell_dir / fname
            if not path.exists():
                continue
            df_eis = pd.read_csv(path)
            df_eis = df_eis.dropna(subset=["Frequency", "ReZ"]).sort_values("Frequency")
            if len(df_eis) < 2:
                continue
            r_diff = float(df_eis.iloc[0]["ReZ"])    # lowest freq
            r_ohmic = float(df_eis.iloc[-1]["ReZ"])  # highest freq
            rows.append({"cell_id": cell, "day": day, "R_ohmic": r_ohmic, "R_diff": r_diff})
    return pd.DataFrame(rows)


def main():
    print("=== Extracting Khan 2025 features ===")
    cap_df = extract_qmax_per_cell_per_day()
    eis_df = extract_eis_per_cell_per_day(soc="S50")
    print(f"Capacity observations: {len(cap_df)} from {cap_df['cell_id'].nunique()} cells")
    print(f"EIS observations:      {len(eis_df)} from {eis_df['cell_id'].nunique()} cells")

    df = cap_df.merge(eis_df, on=["cell_id", "day"], how="inner")
    print(f"Joined (capacity + EIS): {len(df)} from {df['cell_id'].nunique()} cells\n")

    operators = ["Q_max_Ah", "R_ohmic", "R_diff"]
    # Per-cell standardization using day-0
    cell_stats = {}
    z_rows = []
    fresh_pool = []
    for cell, group in df.groupby("cell_id"):
        group = group.sort_values("day").reset_index(drop=True)
        fresh = group[group["day"] == 0]
        if fresh.empty:
            continue
        # Day 0 has 1 obs per cell; we'll pool across cells for cov + use cohort-mean for sd
        cell_stats[cell] = fresh[operators].values[0]

    if not cell_stats:
        print("[abort] no day-0 fresh references available")
        return

    # Stage 1: standardize using per-cell day-0 as mu; sd comes from pooled across-cell day-0 variance
    day0_stack = np.vstack(list(cell_stats.values()))
    pooled_sd = day0_stack.std(axis=0, ddof=1)
    pooled_sd = np.where(pooled_sd < 1e-12, 1e-12, pooled_sd)
    print(f"Pooled day-0 SD (across {len(cell_stats)} cells):")
    for k, op in enumerate(operators):
        print(f"  {op}: {pooled_sd[k]:.4f}")

    for cell, group in df.groupby("cell_id"):
        group = group.sort_values("day").reset_index(drop=True)
        if cell not in cell_stats:
            continue
        mu = cell_stats[cell]
        z = (group[operators].values - mu) / pooled_sd
        zdf = group[["cell_id", "day"]].copy().reset_index(drop=True)
        for k, op in enumerate(operators):
            zdf[f"z_{op}"] = z[:, k]
        z_rows.append(zdf)
        # Fresh pool = day-0 z-scores (which are all zero by construction → can't use)
        # Instead, use cross-cell day-0 dispersion as the empirical fresh cov
    z_all = pd.concat(z_rows, ignore_index=True)
    z_cols = [f"z_{op}" for op in operators]

    # Fresh cov = empirical covariance of day-0 standardized residuals (across cells)
    day0_z = z_all[z_all["day"] == 0][z_cols].values
    # Subtract the mean (should be zero by construction); compute cov
    pooled_cov = np.cov(day0_z.T, ddof=1) if len(day0_z) > 1 else np.eye(3)
    # Regularize
    pooled_cov = pooled_cov + 1e-3 * np.eye(3)
    cov_inv = np.linalg.inv(pooled_cov)

    # Mahalanobis distance per (cell, day)
    z_mat = z_all[z_cols].values
    d2 = np.einsum("ij,jk,ik->i", z_mat, cov_inv, z_mat)
    z_all["m_dist"] = np.sqrt(d2)

    print("\n=== PPC: day-0 dist^2 should be small (cells are mean-centered) ===")
    fresh_d2 = d2[z_all["day"].values == 0]
    print(f"  day-0 n={len(fresh_d2)}, mean d^2={np.mean(fresh_d2):.3f}")

    thr = np.sqrt(stats.chi2.ppf(0.99, df=3))
    print(f"\nThreshold (99th pct chi^2(3)): {thr:.2f}")

    print("\n=== Mahalanobis trajectory per cell ===")
    pivot = z_all.pivot(index="day", columns="cell_id", values="m_dist")
    print(pivot.to_string(float_format=lambda x: f"{x:6.2f}" if pd.notna(x) else "    -"))

    # === Phase 4 classifier ===
    print("\n=== Phase 4 classifier: per-cell residual-direction unit vector ===")
    classifications = []
    for cell, group in z_all.groupby("cell_id"):
        group = group.sort_values("day").reset_index(drop=True)
        flagged = group[group["m_dist"] > thr]
        if flagged.empty:
            classifications.append({"cell_id": cell, "n_flagged": 0, "class": "unflagged", "confidence": 0.0,
                                    "s_LLI": np.nan, "s_LAM_SEI": np.nan})
            continue
        z_unit = flagged[z_cols].values
        norms = np.linalg.norm(z_unit, axis=1, keepdims=True)
        u = z_unit / np.where(norms < 1e-12, 1e-12, norms)
        u_median = np.median(u, axis=0)
        # Re-normalize median (medians don't preserve unit length)
        u_median = u_median / max(np.linalg.norm(u_median), 1e-12)
        s_lli = float(np.dot(u_median, U_LLI))
        s_lam = float(np.dot(u_median, U_LAM_SEI))
        conf = abs(s_lli - s_lam)
        if conf < CONF_THRESHOLD:
            cls = "unclassified"
        else:
            cls = "LLI" if s_lli > s_lam else "LAM+SEI"
        classifications.append({"cell_id": cell, "n_flagged": int(len(flagged)), "class": cls,
                                "confidence": conf, "s_LLI": s_lli, "s_LAM_SEI": s_lam})

    cdf = pd.DataFrame(classifications)
    print(cdf.to_string(index=False, float_format=lambda x: f"{x:.3f}" if isinstance(x, float) else str(x)))

    # === Falsification checks ===
    n_cells = len(cdf)
    n_classified = (cdf["class"].isin(["LLI", "LAM+SEI"])).sum()
    n_lam = (cdf["class"] == "LAM+SEI").sum()
    n_lli = (cdf["class"] == "LLI").sum()

    print("\n=== Pre-registered falsification checks ===")
    print(f"  Total cells: {n_cells}")
    print(f"  Confidently classified (confidence >= {CONF_THRESHOLD}): {n_classified} ({n_classified/max(n_cells,1)*100:.1f}%)  -> pre-reg requires >=50%")
    if n_classified > 0:
        lam_frac = n_lam / n_classified
        print(f"  LAM+SEI among classified: {n_lam}/{n_classified} ({lam_frac*100:.1f}%) -> pre-reg expects >=70% for NMC")
        print(f"  LLI among classified: {n_lli}/{n_classified} ({n_lli/n_classified*100:.1f}%)")

    # Permutation null: shuffle assignment of trajectories to centroids
    rng = np.random.default_rng(seed=42)
    observed_separation = abs(cdf[cdf["class"]=="LLI"]["confidence"].mean() if n_lli else 0
                              - cdf[cdf["class"]=="LAM+SEI"]["confidence"].mean() if n_lam else 0)
    # Simpler null: how often does a random class-assignment produce as many confident classifications?
    null_n_confident = []
    for _ in range(10000):
        shuffled_confidences = rng.permutation(cdf["confidence"].values)
        null_n_confident.append(int((shuffled_confidences >= CONF_THRESHOLD).sum()))
    null_n_confident = np.array(null_n_confident)
    obs_conf = int((cdf["confidence"] >= CONF_THRESHOLD).sum())
    perm_p = float((null_n_confident >= obs_conf).mean())
    print(f"\n  Permutation null (10000 shuffles of confidence): p = {perm_p:.4f}")
    print(f"  Bonferroni threshold for 3 replication cohorts: alpha/3 = 0.0167")

    cdf.to_parquet(OUT_DIR / "khan_2025_classification.parquet")
    print(f"\nWritten: {OUT_DIR / 'khan_2025_classification.parquet'}")


if __name__ == "__main__":
    main()
