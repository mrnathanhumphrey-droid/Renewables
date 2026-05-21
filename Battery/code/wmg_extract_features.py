"""
WMG 25-cell (Rashid 2023) extraction → per-cell residual-direction vector.

WMG design = cross-sectional aging. Each cell terminates at one of {80, 85, 90, 95, 100}% SOH.
EIS is measured at that one terminal SOH stage at 5 SOC x 3 T = 15 configs per cell.
Capacity vs cycle is tracked throughout for each cell.

For the C1 cross-chemistry hierarchical model, we want one (z_Q_max, z_R_ohmic, z_R_diff)
observation per cell, matching the operator triad used in Khan / SECL canonical (SOC=50, T=25C).

Fresh reference: the 5 control cells (Cell28-32) at 100SOH, SOC=50, T=25C.
Aged cells (19): contribute one terminal residual vector each.
"""

from pathlib import Path
import re
import numpy as np
import pandas as pd
import openpyxl


ROOT = Path("D:/Renewables/Battery/data/wmg_25cell/DIB_Data/DIB_Data")
EIS_DIR = ROOT / ".csvfiles" / "EIS_Test"
CAP_XLSX = ROOT / "CapacityVsCycleNumber.xlsx"
OUT_DIR = Path("D:/Renewables/Battery/data/processed")

EIS_PATTERN = re.compile(r"^Cell(\d+)_(\d+)SOH_(\d+)degC_(\d+)SOC_\d+\.xls$")

# Canonical operator-extraction config (matches Khan)
CANONICAL_SOC = 50
CANONICAL_T = 25


def parse_eis_files():
    """Walk EIS_Test/, parse filenames, return DataFrame [cell, soh, t, soc, path]."""
    rows = []
    for path in sorted(EIS_DIR.glob("*.xls")):
        m = EIS_PATTERN.match(path.name)
        if not m:
            continue
        cell = int(m.group(1))
        soh = int(m.group(2))
        t = int(m.group(3))
        soc = int(m.group(4))
        rows.append({"cell": cell, "soh": soh, "t": t, "soc": soc, "path": str(path)})
    return pd.DataFrame(rows)


def extract_eis_features(xls_path):
    """Return (R_ohmic, R_diff) from a single EIS xls file.
    File has no header: cols = [freq_Hz, ReZ, -ImZ], 60 rows desc freq.
    R_ohmic = ReZ at highest freq; R_diff = ReZ at lowest freq.
    """
    df = pd.read_excel(xls_path, header=None)
    df.columns = ["freq", "ReZ", "negImZ"]
    df = df.sort_values("freq")
    r_diff = float(df.iloc[0]["ReZ"])    # lowest freq
    r_ohmic = float(df.iloc[-1]["ReZ"])  # highest freq
    return r_ohmic, r_diff


def load_capacity_long():
    """Parse CapacityVsCycleNumber.xlsx into long DataFrame [cell, cycle, capacity_Ah, soh_pct]."""
    wb = openpyxl.load_workbook(CAP_XLSX, data_only=True)
    rows = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) < 3:
            continue
        header = all_rows[0]   # CN, Cell X, None, Cell Y, None, ...
        # subheader at row 1: None, Capacity (Ah), SOH, Capacity, SOH, ...
        # Build cell index → (cap_col, soh_col)
        cell_cols = []
        for i, h in enumerate(header):
            if h is None or not isinstance(h, str) or not h.startswith("Cell"):
                continue
            m = re.match(r"^Cell\s*(\d+)$", h.strip())
            if m:
                cell_cols.append((int(m.group(1)), i, i + 1))  # cap=i, soh=i+1
        for r in all_rows[2:]:
            cn = r[0]
            if cn is None or not isinstance(cn, (int, float)):
                continue
            for cell, cap_col, soh_col in cell_cols:
                cap = r[cap_col] if cap_col < len(r) else None
                soh = r[soh_col] if soh_col < len(r) else None
                if cap is None or soh is None:
                    continue
                try:
                    rows.append({"cell": cell, "cycle": int(cn),
                                 "capacity_Ah": float(cap), "soh_pct": float(soh)})
                except (TypeError, ValueError):
                    continue
    return pd.DataFrame(rows)


def main():
    print("=== WMG 25-cell extraction ===\n")

    # 1. EIS file inventory
    eis_inv = parse_eis_files()
    print(f"EIS file inventory: {len(eis_inv)} files across {eis_inv['cell'].nunique()} cells")
    print(f"  Cells x SOH:\n{eis_inv.groupby(['soh']).agg(n_cells=('cell','nunique'), n_files=('path','count'))}\n")

    # 2. Filter EIS to canonical config (SOC=50, T=25)
    eis_canon = eis_inv[(eis_inv["soc"] == CANONICAL_SOC) & (eis_inv["t"] == CANONICAL_T)].copy()
    print(f"Canonical EIS (SOC={CANONICAL_SOC}, T={CANONICAL_T}C): {len(eis_canon)} files (one per cell)\n")

    # 3. Extract R_ohmic + R_diff per cell
    feats = []
    for _, r in eis_canon.iterrows():
        try:
            r_oh, r_di = extract_eis_features(r["path"])
            feats.append({"cell": int(r["cell"]), "soh_eis": int(r["soh"]),
                          "R_ohmic": r_oh, "R_diff": r_di})
        except Exception as e:
            print(f"  [err] {r['path']}: {e}")
    feat_df = pd.DataFrame(feats)
    print("Per-cell EIS features:")
    print(feat_df.sort_values(["soh_eis", "cell"]).to_string(index=False,
        float_format=lambda x: f"{x:.5f}" if isinstance(x, float) else str(x)))

    # 4. Load capacity → derive Q_max per cell at EIS SOH point
    cap_long = load_capacity_long()
    print(f"\nCapacity long table: {len(cap_long)} rows, {cap_long['cell'].nunique()} unique cells")

    # For each cell, get fresh Q_max (cycle 0) and Q_max at its EIS SOH stage
    qrows = []
    for cell in feat_df["cell"].unique():
        soh_target = int(feat_df[feat_df["cell"] == cell]["soh_eis"].iloc[0])
        c_traj = cap_long[cap_long["cell"] == cell].sort_values("cycle")
        if c_traj.empty:
            qrows.append({"cell": cell, "Q_fresh": np.nan, "Q_at_eis_soh": np.nan, "cycle_at_eis_soh": np.nan})
            continue
        q_fresh = float(c_traj[c_traj["cycle"] == 0]["capacity_Ah"].iloc[0]) if (c_traj["cycle"] == 0).any() else np.nan
        if soh_target == 100:
            # Control cell: only cycle 0 measurement
            q_at = q_fresh
            cyc_at = 0
        else:
            # Find cycle where SOH closest to target (and capacity actually present)
            non0 = c_traj[c_traj["cycle"] > 0]
            if non0.empty:
                q_at = np.nan
                cyc_at = np.nan
            else:
                # closest absolute SOH to target
                idx = (non0["soh_pct"] - soh_target).abs().idxmin()
                q_at = float(non0.loc[idx, "capacity_Ah"])
                cyc_at = int(non0.loc[idx, "cycle"])
        qrows.append({"cell": cell, "Q_fresh": q_fresh, "Q_at_eis_soh": q_at, "cycle_at_eis_soh": cyc_at})

    qdf = pd.DataFrame(qrows)

    # 5. Merge and compute per-cell residual triad
    full = feat_df.merge(qdf, on="cell").sort_values(["soh_eis", "cell"]).reset_index(drop=True)
    full["Q_max_Ah"] = full["Q_at_eis_soh"]

    print("\nPer-cell joined features:")
    print(full[["cell", "soh_eis", "cycle_at_eis_soh", "Q_fresh", "Q_max_Ah", "R_ohmic", "R_diff"]].to_string(
        index=False, float_format=lambda x: f"{x:.4f}" if isinstance(x, float) else str(x)))

    # 6. Fresh-cell pool (100SOH controls) → mean & SD of operators across the 5 controls
    fresh = full[full["soh_eis"] == 100].dropna(subset=["Q_max_Ah", "R_ohmic", "R_diff"])
    aged = full[full["soh_eis"] != 100].dropna(subset=["Q_max_Ah", "R_ohmic", "R_diff"])

    print(f"\nFresh-pool cells (100SOH controls): n={len(fresh)}")
    print(f"Aged cells:                           n={len(aged)}")

    fresh_means = fresh[["Q_max_Ah", "R_ohmic", "R_diff"]].mean()
    fresh_sds = fresh[["Q_max_Ah", "R_ohmic", "R_diff"]].std(ddof=1)
    print(f"\nFresh-pool means: Q={fresh_means['Q_max_Ah']:.4f}  R_oh={fresh_means['R_ohmic']:.5f}  R_di={fresh_means['R_diff']:.5f}")
    print(f"Fresh-pool SDs:   Q={fresh_sds['Q_max_Ah']:.4f}  R_oh={fresh_sds['R_ohmic']:.5f}  R_di={fresh_sds['R_diff']:.5f}")

    # 7. Per-aged-cell residual triad
    # Note: lower Q_max = degradation = NEGATIVE in delta-Q direction.
    # Higher R_ohmic / R_diff = degradation = POSITIVE.
    # Centroids: U_LLI = (-1, 0, 0); U_LAM_SEI = (-1, +1, +1)/sqrt(3)
    aged_z = aged.copy()
    aged_z["z_Q_max"] = (aged["Q_max_Ah"] - fresh_means["Q_max_Ah"]) / fresh_sds["Q_max_Ah"]
    aged_z["z_R_ohmic"] = (aged["R_ohmic"] - fresh_means["R_ohmic"]) / fresh_sds["R_ohmic"]
    aged_z["z_R_diff"] = (aged["R_diff"] - fresh_means["R_diff"]) / fresh_sds["R_diff"]

    # Unit residual direction per cell
    z_mat = aged_z[["z_Q_max", "z_R_ohmic", "z_R_diff"]].values
    norms = np.linalg.norm(z_mat, axis=1, keepdims=True)
    u_mat = z_mat / np.where(norms < 1e-12, 1e-12, norms)
    aged_z["u_Q_max"] = u_mat[:, 0]
    aged_z["u_R_ohmic"] = u_mat[:, 1]
    aged_z["u_R_diff"] = u_mat[:, 2]
    aged_z["mahal"] = norms.flatten()  # std-normal-equivalent distance (diagonal cov)

    print("\n=== Per-cell standardized triad + unit direction ===")
    print(aged_z[["cell", "soh_eis", "z_Q_max", "z_R_ohmic", "z_R_diff", "u_Q_max", "u_R_ohmic", "u_R_diff", "mahal"]].to_string(
        index=False, float_format=lambda x: f"{x:+.3f}" if isinstance(x, float) else str(x)))

    # 8. Save
    out_path = OUT_DIR / "wmg_25cell_classification.parquet"
    aged_z["chemistry"] = "NMC811"
    aged_z["form_factor"] = "cyl"
    aged_z["cohort"] = "WMG"
    aged_z.to_parquet(out_path)
    print(f"\nWritten: {out_path}")

    # Quick centroid-distance preview vs pre-reg
    U_LLI = np.array([-1.0, 0.0, 0.0])
    U_LAM_SEI = np.array([-1.0, +1.0, +1.0]) / np.sqrt(3.0)
    print("\n=== Preview vs Phase-4 pre-reg centroids ===")
    print(f"{'cell':>5} {'soh':>4} {'sim_LLI':>8} {'sim_LAM':>8} {'argmax':>8} {'conf':>6}")
    for _, r in aged_z.iterrows():
        u = np.array([r["u_Q_max"], r["u_R_ohmic"], r["u_R_diff"]])
        s_lli = float(u @ U_LLI)
        s_lam = float(u @ U_LAM_SEI)
        conf = abs(s_lli - s_lam)
        argmax = "LLI" if s_lli > s_lam else "LAM+SEI"
        print(f"{int(r['cell']):>5} {int(r['soh_eis']):>4} {s_lli:+.3f}  {s_lam:+.3f}    {argmax:>8} {conf:.3f}")


if __name__ == "__main__":
    main()
