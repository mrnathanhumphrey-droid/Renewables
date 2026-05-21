"""
Phase 2.3 (cont'd) — second-life feature extraction.

Builds per-(cell, RPT) features for the 6 second-life cells.

Features:
  Q_max_Ah:                            from raw Channel_*.xlsx files
  R_ohmic_v326, R_ohmic_v363, R_ohmic_v400:   from EIS *_ALL.xlsx (RPTs 5-8)
                                              + eis_N.mat where present (RPT 7)
  R_diff_v326,  R_diff_v363,  R_diff_v400:    same source as R_ohmic
  (voltage 326/363/400 mV correspond to low/mid/high SOC)

RPTs 9-19 EIS pulled from per-cell folders is a TODO — those use raw Gamry
.DTA files inside per-cell subfolders. Deferred to next push.

Outputs: data/processed/features_second_life.parquet
"""

from pathlib import Path
import re
import numpy as np
import pandas as pd
import openpyxl
import scipy.io as sio


BASE = Path("D:/Renewables/Battery/data/secl_second_life/SL_Dataset_SECL_INR21700-M50T")
OUT_DIR = Path("D:/Renewables/Battery/data/processed")
CELLS = ["G1", "V4", "V5", "W8", "W9", "W10"]


def extract_capacity_max(xlsx_path: Path) -> float:
    """Open a capacity-test .xlsx and return max-min discharge capacity (Ah)."""
    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    except Exception as exc:
        print(f"  [load fail] {xlsx_path.name}: {exc}")
        return float("nan")
    ch_sheets = [s for s in wb.sheetnames if s.startswith("Channel_")]
    if not ch_sheets:
        return float("nan")
    ws = wb[ch_sheets[0]]
    # Discharge_Capacity column = index 8 (0-based). Iterate rows.
    max_q = -np.inf
    min_q = np.inf
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) > 8:
            q = row[8]
            if q is not None:
                if q > max_q: max_q = q
                if q < min_q: min_q = q
    if max_q == -np.inf:
        return float("nan")
    return float(max_q - min_q)


def scan_capacity(rpt_n: int):
    """Find capacity xlsx files per cell at the given RPT. Returns {cell: path}."""
    rpt_dir = BASE / "diagnostic_tests" / f"RPT_{rpt_n}" / "Capacity_test_with_pulses"
    if not rpt_dir.exists():
        return {}
    out = {}
    for cell_dir in rpt_dir.iterdir():
        if not cell_dir.is_dir():
            continue
        # folder name like 20230322_g1_RPT  → cell = g1
        m = re.match(r"^\d{8}_([a-z0-9]+)_RPT$", cell_dir.name)
        if not m:
            continue
        cell = m.group(1).upper()
        if cell not in CELLS:
            continue
        # find the channel xlsx
        for ch_dir in cell_dir.iterdir():
            if ch_dir.is_dir() and ch_dir.name.startswith("Channel_"):
                xlsx_files = list(ch_dir.glob("*.xlsx"))
                if xlsx_files:
                    out[cell] = xlsx_files[0]
                    break
    return out


def parse_eis_all_xlsx(xlsx_path: Path, voltage_label: str):
    """Parse a *_ALL.xlsx EIS file. Returns {cell: (R_ohmic, R_diff)} for the file's voltage.

    Each pair of columns is (Re_Z, Im_Z) for one cell. Header row identifies cell from filename
    embedded in the column title 'ZCURVE (YYYYMMDD_{CELL}_EIS{V}_Chan...)'.
    """
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    # column → cell mapping
    col_cell = {}
    for c_idx, hdr in enumerate(header):
        if hdr and "ZCURVE" in str(hdr):
            m = re.search(r"_([A-Za-z0-9]+)_EIS", str(hdr))
            if m:
                col_cell[c_idx] = m.group(1).upper()
    # Collect all rows into 2D array
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(row)
    if not data:
        return {}
    arr = np.array([[(v if v is not None else np.nan) for v in r] for r in data], dtype=float)
    n_rows = arr.shape[0]

    # IMPORTANT: row order in these files is from HIGH freq to LOW freq (verified vs first-life
    # 10 kHz on row 1). Re_Z and Im_Z are adjacent columns: Re at c, Im at c+1.
    out = {}
    for c_idx, cell in col_cell.items():
        re_col = arr[:, c_idx]
        # first valid row = R_ohmic (high freq), last valid row = R_diff (low freq)
        valid = ~np.isnan(re_col)
        if not valid.any():
            continue
        first_valid = np.argmax(valid)  # index of first True
        last_valid = n_rows - 1 - np.argmax(valid[::-1])
        out[cell] = (float(re_col[first_valid]), float(re_col[last_valid]))
    return out


def scan_eis_all(rpt_n: int):
    """Find EIS _ALL.xlsx files at RPT_N for the three voltages. Returns {voltage: path}."""
    rpt_dir = BASE / "diagnostic_tests" / f"RPT_{rpt_n}"
    # _ALL files may be at RPT_N level (RPT_7) or inside EIS_N/ subdir (RPT_5,6,8)
    candidates = list(rpt_dir.glob("*_EIS_*_ALL.xlsx"))
    if not candidates:
        eis_dir = rpt_dir / f"EIS_{rpt_n}"
        if eis_dir.exists():
            candidates = list(eis_dir.glob("*_EIS_*_ALL.xlsx"))
    out = {}
    for p in candidates:
        m = re.search(r"_(\d{3})V_ALL", p.name)
        if m:
            out[m.group(1)] = p
    return out


def extract_eis_from_mat(rpt_n: int):
    """Use eis_N.mat (only RPT 7 currently) for EIS features per cell."""
    mat_path = BASE / "diagnostic_tests" / f"RPT_{rpt_n}" / f"eis_{rpt_n}.mat"
    if not mat_path.exists():
        return None
    m = sio.loadmat(str(mat_path), squeeze_me=True)
    cell_order = [str(x).strip() for x in m["cn_fn_chan_order"]]
    # voltage labels are reverse of file naming convention; canonicalize
    eis_volts_field = m["EIS_volts"]
    vlabels = [str(v).strip() for v in eis_volts_field]
    out = {}
    for v in vlabels:
        key = f"EIS_{rpt_n}_{v}"
        if key not in m:
            continue
        arr = m[key]  # (n_freqs, 2*n_cells)
        # pairs (Re, Im) per cell in cell_order
        n_cells = len(cell_order)
        for i, cell in enumerate(cell_order):
            re_col = arr[:, 2 * i]
            # row 0 = high freq, row -1 = low freq
            r_ohmic = float(re_col[0])
            r_diff = float(re_col[-1])
            out.setdefault(cell, {})[v] = (r_ohmic, r_diff)
    return out


def main():
    rows = []
    n_rpts = 19
    for rpt_idx in range(1, n_rpts + 1):
        # capacity
        cap_paths = scan_capacity(rpt_idx)
        # EIS via _ALL.xlsx
        eis_v_paths = scan_eis_all(rpt_idx)
        # also via eis_N.mat
        eis_mat = extract_eis_from_mat(rpt_idx)

        print(f"RPT_{rpt_idx}: cap files={len(cap_paths)} cells, EIS _ALL={list(eis_v_paths.keys())}, mat={'yes' if eis_mat else 'no'}")

        eis_xlsx_data = {}  # voltage -> {cell: (R_ohmic, R_diff)}
        for v, p in eis_v_paths.items():
            try:
                eis_xlsx_data[v] = parse_eis_all_xlsx(p, v)
            except Exception as exc:
                print(f"  [eis parse fail] RPT_{rpt_idx} v={v}: {exc}")

        for cell in CELLS:
            row = {
                "cell_id": cell,
                "lifecycle": "second_life",
                "rpt_idx": rpt_idx,
                "Q_max_Ah": float("nan"),
            }
            for v in ("326", "363", "400"):
                row[f"R_ohmic_v{v}"] = float("nan")
                row[f"R_diff_v{v}"] = float("nan")

            if cell in cap_paths:
                row["Q_max_Ah"] = extract_capacity_max(cap_paths[cell])

            # prefer .mat if present
            if eis_mat and cell in eis_mat:
                for v, (r_o, r_d) in eis_mat[cell].items():
                    row[f"R_ohmic_v{v}"] = r_o
                    row[f"R_diff_v{v}"] = r_d
            else:
                for v, cell_map in eis_xlsx_data.items():
                    if cell in cell_map:
                        r_o, r_d = cell_map[cell]
                        row[f"R_ohmic_v{v}"] = r_o
                        row[f"R_diff_v{v}"] = r_d

            rows.append(row)

    df = pd.DataFrame(rows)
    print("\n=== Second-life feature availability ===")
    summary = df.groupby("cell_id").agg(
        rpts=("rpt_idx", "count"),
        cap_n=("Q_max_Ah", lambda s: s.notna().sum()),
        eis_ohmic_363_n=("R_ohmic_v363", lambda s: s.notna().sum()),
        eis_diff_363_n=("R_diff_v363", lambda s: s.notna().sum()),
    )
    print(summary.to_string())

    print("\n=== Q_max trajectory ===")
    pivot = df.pivot(index="rpt_idx", columns="cell_id", values="Q_max_Ah")
    print(pivot.to_string(float_format=lambda x: f"{x:.3f}" if pd.notna(x) else "    -"))

    out_path = OUT_DIR / "features_second_life.parquet"
    df.to_parquet(out_path)
    print(f"\nWritten: {out_path}")


if __name__ == "__main__":
    main()
