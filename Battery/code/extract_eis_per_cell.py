"""
Phase 2.3 (cont'd) — second-life per-cell EIS extraction for RPTs 9-19.

Folder structure:
  RPT_N/EIS_N/YYYYMMDD_{cell}_EIS{voltage}V/Channel_K/YYYYMMDD_{cell}_EIS{voltage}V_Channel_K.xlsx

Each .xlsx has an ACIM_Chan* sheet with columns:
  Device_ID, Test_ID, Channel_ID, Cycle_ID, Step_ID, Pt, Freq, Zmod, Zphz

Extracted features (per cell, per RPT, per voltage):
  R_ohmic_v{V}: Re(Z) at highest frequency (Pt=0, ~10 kHz)
  R_diff_v{V}:  Re(Z) at lowest frequency (Pt=42, ~10 mHz)

Re(Z) = Zmod * cos(Zphz_rad)

Output: merged into data/processed/features_second_life.parquet, replacing
EIS columns where new data is available.
"""

from pathlib import Path
import re
import math
import numpy as np
import pandas as pd
import openpyxl


BASE = Path("D:/Renewables/Battery/data/secl_second_life/SL_Dataset_SECL_INR21700-M50T/diagnostic_tests")
OUT_DIR = Path("D:/Renewables/Battery/data/processed")
CELLS = ["G1", "V4", "V5", "W8", "W9", "W10"]


def parse_acim_sheet(xlsx_path: Path):
    """Return (R_ohmic, R_diff) from ACIM_Chan* sheet, or (nan, nan)."""
    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    except Exception:
        return float("nan"), float("nan")
    acim_sheets = [s for s in wb.sheetnames if s.startswith("ACIM_")]
    if not acim_sheets:
        return float("nan"), float("nan")
    ws = wb[acim_sheets[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    if not rows:
        return float("nan"), float("nan")
    # Columns: Device_ID(0), Test_ID(1), Channel_ID(2), Cycle_ID(3), Step_ID(4), Pt(5), Freq(6), Zmod(7), Zphz(8)
    points = []
    for r in rows:
        if len(r) < 9:
            continue
        try:
            freq = float(r[6])
            zmod = float(r[7])
            zphz_deg = float(r[8])
            re_z = zmod * math.cos(math.radians(zphz_deg))
            points.append((freq, re_z))
        except (TypeError, ValueError):
            continue
    if not points:
        return float("nan"), float("nan")
    # sort by freq ascending so lowest is first, highest last
    points.sort(key=lambda x: x[0])
    r_diff = points[0][1]   # lowest freq
    r_ohmic = points[-1][1] # highest freq
    return float(r_ohmic), float(r_diff)


def find_cell_folders(eis_dir: Path):
    """Return {(cell, voltage): folder_path} for per-cell EIS subfolders."""
    out = {}
    if not eis_dir.exists():
        return out
    for child in eis_dir.iterdir():
        if not child.is_dir():
            continue
        m = re.match(r"^\d{8}_([A-Za-z0-9]+)_EIS(\d{3})V?$", child.name)
        if not m:
            continue
        cell = m.group(1).upper()
        voltage = m.group(2)
        if cell not in CELLS:
            continue
        out[(cell, voltage)] = child
    return out


def parse_cell_folder(folder: Path):
    """Find the Channel_K xlsx and return (R_ohmic, R_diff)."""
    for ch_dir in folder.iterdir():
        if ch_dir.is_dir() and ch_dir.name.startswith("Channel_"):
            xlsx_files = list(ch_dir.glob("*.xlsx"))
            if xlsx_files:
                return parse_acim_sheet(xlsx_files[0])
    return float("nan"), float("nan")


def main():
    # Load existing second-life features
    sl_path = OUT_DIR / "features_second_life.parquet"
    df = pd.read_parquet(sl_path)
    print(f"Loaded {len(df)} second-life rows from {sl_path}\n")

    # Update for RPTs 9-19 (also re-do 5-8 for consistency, since per-cell often differs from _ALL slightly)
    update_count = 0
    for rpt_idx in range(5, 20):
        eis_dir = BASE / f"RPT_{rpt_idx}" / f"EIS_{rpt_idx}"
        folder_map = find_cell_folders(eis_dir)
        if not folder_map:
            continue
        print(f"RPT_{rpt_idx}: {len(folder_map)} cell-voltage folders")
        for (cell, voltage), folder in folder_map.items():
            r_ohmic, r_diff = parse_cell_folder(folder)
            if np.isnan(r_ohmic) and np.isnan(r_diff):
                continue
            mask = (df["cell_id"] == cell) & (df["rpt_idx"] == rpt_idx)
            col_ohmic = f"R_ohmic_v{voltage}"
            col_diff = f"R_diff_v{voltage}"
            if col_ohmic in df.columns and col_diff in df.columns:
                df.loc[mask, col_ohmic] = r_ohmic
                df.loc[mask, col_diff] = r_diff
                update_count += 1

    print(f"\nTotal (cell, rpt, voltage) entries written: {update_count}")

    # Coverage report
    print("\n=== Updated second-life EIS coverage (v363 = mid SOC) ===")
    summary = df.groupby("cell_id").agg(
        rpts=("rpt_idx", "count"),
        cap_n=("Q_max_Ah", lambda s: s.notna().sum()),
        eis_ohmic_363_n=("R_ohmic_v363", lambda s: s.notna().sum()),
        eis_diff_363_n=("R_diff_v363", lambda s: s.notna().sum()),
    )
    print(summary.to_string())

    print("\n=== R_ohmic_v363 trajectory per cell ===")
    pivot = df.pivot(index="rpt_idx", columns="cell_id", values="R_ohmic_v363")
    print(pivot.to_string(float_format=lambda x: f"{x:.4f}" if pd.notna(x) else "    -"))

    df.to_parquet(sl_path)
    print(f"\nUpdated: {sl_path}")


if __name__ == "__main__":
    main()
