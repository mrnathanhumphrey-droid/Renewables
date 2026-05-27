"""C3 Khan cross-substrate validation — amendment deployment gate.

Pre-reg: literature/50_khan_cross_substrate_prereg.md (lock a49c830).

Tests whether the locked C3 amendment architecture (variant iv 6D
fresh+aged stack -> z-score -> PCA k=2 -> unit-vector cosine PERMANOVA)
transfers off the PyBaMM synthetic substrate to real-cell Khan 2025
NMC_prism cohort at Khan's natural EIS instrumentation noise.

Pre-reg deviation log (§8 of literature/50):
  Deviation: EIS SOC bin is S50 (mid-SOC), not "high SOC >=90%" per the
    pre-reg's §2 wording. Trigger: existing Khan extraction convention
    (khan_extract_and_classify.py) uses S50; departing would require
    new file-availability validation across all 14 cycle cells. Filename
    convention is `ACR_t25_S50_{day}d_{cell}_convert.csv` (not the ACRS
    prefix seen in some subdirs). Diagnostic-driven, not result-driven.

Operator extraction (matches khan_extract_and_classify.py):
  Q_max  = max Cdcg from CD_t25_{day}d.xlsx sheet {cell}, col 6 from row 7
  R_ohmic = highest-frequency ReZ in EIS spectrum (>= 7 kHz typically)
  R_diff = lowest-frequency ReZ in EIS spectrum (0.01 Hz typically)

Cohort: 14 cycle cells. Day 0 = fresh; Day 90 = aged. Drops:
  - S2, S18 (incomplete-aging exclusion)
  - 5 calendar cells (aging_type='calendar') -- no operating-cycling design
"""

from pathlib import Path
import re
import sys
import numpy as np
import pandas as pd
import openpyxl


DATA = Path("D:/Renewables/Battery/data/khan_2025")
OUT_DIR = Path("D:/Renewables/Battery/data/processed")
EXCLUDE_CELLS = {"S2", "S18"}

FRESH_DAY = 0
AGED_DAY = 90
EIS_SOC = "S50"

N_PERMS = 10000
ALPHA_BONFERRONI = 0.0167
EFFECT_F_FLOOR = 3.0
EFFECT_F_WEAK = 2.0
ALPHA_WEAK = 0.05

# F4 predictive-transfer parameters
N_HOLDOUT = 4
N_HOLDOUT_REPS = 100
F4_ACC_FLOOR = 0.50  # > chance + 0.17 for 3-level params; > chance + 0.0 for 2-level


def extract_qmax(cells, days):
    rows = []
    for day in days:
        path = DATA / "capacity" / f"CD_t25_{day}d.xlsx"
        if not path.exists():
            print(f"  WARN: capacity file missing: {path.name}")
            continue
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        for sn in wb.sheetnames:
            if not re.match(r"^S\d+$", sn):
                continue
            if sn in EXCLUDE_CELLS or sn not in cells:
                continue
            ws = wb[sn]
            max_cdcg = -np.inf
            for row in ws.iter_rows(min_row=7, values_only=True):
                if len(row) > 6:
                    v = row[6]
                    if v is not None and isinstance(v, (int, float)) and v > max_cdcg:
                        max_cdcg = v
            if max_cdcg != -np.inf:
                rows.append({"cell_id": sn, "day": day, "Q_max": float(max_cdcg)})
    return pd.DataFrame(rows)


def extract_eis(cells, days, soc=EIS_SOC):
    rows = []
    eis_root = DATA / "eis_csv"
    for cell in cells:
        cell_dir = eis_root / cell
        if not cell_dir.exists():
            print(f"  WARN: EIS dir missing for cell {cell}")
            continue
        for day in days:
            fname = f"ACR_t25_{soc}_{day}d_{cell}_convert.csv"
            path = cell_dir / fname
            if not path.exists():
                # Try ACRS prefix as fallback
                fname_alt = f"ACRS_t25_{soc}_{day}d_{cell}_convert.csv"
                path = cell_dir / fname_alt
                if not path.exists():
                    # SOC subdir variants (S100 2 etc.)
                    for socdir in [f"{soc}", f"{soc} 2", f"{soc.lower()}"]:
                        cand = cell_dir / socdir / fname
                        if cand.exists():
                            path = cand
                            break
                        cand = cell_dir / socdir / fname_alt
                        if cand.exists():
                            path = cand
                            break
            if not path.exists():
                print(f"  WARN: EIS file missing: {cell}/day{day} SOC={soc}")
                continue
            df_eis = pd.read_csv(path).dropna(subset=["Frequency", "ReZ"]).sort_values("Frequency")
            if len(df_eis) < 2:
                continue
            r_diff = float(df_eis.iloc[0]["ReZ"])    # lowest freq
            r_ohmic = float(df_eis.iloc[-1]["ReZ"])  # highest freq
            rows.append({"cell_id": cell, "day": day, "R_ohmic": r_ohmic, "R_diff": r_diff})
    return pd.DataFrame(rows)


def build_feature_table():
    """Build per-cell (fresh + aged) operator triplets + design labels."""
    cond = pd.read_csv(DATA / "cell_conditions.csv")
    cycle = cond[(cond["aging_type"] == "cycle") & (~cond["cell"].isin(EXCLUDE_CELLS))].copy()
    cells = sorted(cycle["cell"].tolist())
    print(f"Cycle cells (filtered): {len(cells)} {cells}")

    cap = extract_qmax(cells, [FRESH_DAY, AGED_DAY])
    eis = extract_eis(cells, [FRESH_DAY, AGED_DAY])

    # Pivot to one row per cell
    cap_p = cap.pivot(index="cell_id", columns="day", values="Q_max")
    cap_p.columns = [f"Q_max_d{d}" for d in cap_p.columns]
    eis_p = eis.pivot(index="cell_id", columns="day")
    eis_p.columns = [f"{m}_d{d}" for m, d in eis_p.columns]

    feats = cap_p.join(eis_p)
    feats = feats.dropna()
    feats = feats.reset_index().rename(columns={"cell_id": "cell"})

    # Merge with design labels
    merged = feats.merge(cycle[["cell", "T_C", "soc_range", "charge_rate", "voltage_limits"]], on="cell")
    print(f"\nCells with complete fresh+aged data: {len(merged)}")
    if len(merged) < 10:
        print("  ERROR: too few complete cells for PERMANOVA + F4 holdout")
    return merged


def build_pca2_features(df):
    """variant (iv) 6D stack -> centered z-score -> PCA k=2."""
    cols = ["Q_max_d0", "R_ohmic_d0", "R_diff_d0",
            "Q_max_d90", "R_ohmic_d90", "R_diff_d90"]
    feats = df[cols].values
    feats = feats - feats.mean(axis=0, keepdims=True)
    pooled_sd = feats.std(axis=0, ddof=1)
    pooled_sd = np.where(pooled_sd < 1e-12, 1e-12, pooled_sd)
    z = feats / pooled_sd
    cov = np.cov(z.T)
    eigvals, eigvecs = np.linalg.eigh(cov)
    idx = np.argsort(eigvals)[::-1]
    components = eigvecs[:, idx][:, :2]
    explained = eigvals[idx][:2].sum() / eigvals.sum()
    pca2 = z @ components
    return pca2, components, float(explained), z, cols


def cosine_dist(u):
    cos_mat = u @ u.T
    cos_mat = np.clip(cos_mat, -1.0, 1.0)
    return 1.0 - cos_mat


def unit_project(z):
    n = np.linalg.norm(z, axis=1, keepdims=True)
    return z / np.where(n < 1e-12, 1e-12, n)


def permanova_pseudoF(d_mat, labels):
    n = len(d_mat)
    labels = np.asarray(labels)
    unique = np.unique(labels)
    a = len(unique)
    if a < 2 or n - a < 1:
        return float("nan")
    total_ss = float((d_mat ** 2).sum()) / (2.0 * n)
    within_ss = 0.0
    for g in unique:
        mask = labels == g
        n_g = int(mask.sum())
        if n_g < 2:
            continue
        sub = d_mat[np.ix_(mask, mask)]
        within_ss += float((sub ** 2).sum()) / (2.0 * n_g)
    between_ss = total_ss - within_ss
    if within_ss <= 0 or between_ss <= 0:
        return float("nan")
    return (between_ss / (a - 1)) / (within_ss / (n - a))


def permanova_test(u, labels, rng, n_perms=N_PERMS):
    d_mat = cosine_dist(u)
    F_obs = permanova_pseudoF(d_mat, labels)
    if not np.isfinite(F_obs):
        return float("nan"), float("nan")
    labels = np.asarray(labels)
    n_ge = 0
    for _ in range(n_perms):
        perm = rng.permutation(labels)
        F_p = permanova_pseudoF(d_mat, perm)
        if np.isfinite(F_p) and F_p >= F_obs:
            n_ge += 1
    return F_obs, (n_ge + 1) / (n_perms + 1)


def verdict(F, p):
    if not (np.isfinite(F) and np.isfinite(p)):
        return "NULL"
    if p < ALPHA_BONFERRONI and F > EFFECT_F_FLOOR:
        return "PASS"
    if ALPHA_BONFERRONI <= p < ALPHA_WEAK and F > EFFECT_F_WEAK:
        return "WEAK PASS"
    return "NULL"


def centroid_check(u, labels):
    """P-Khan_F3: pairwise cosines between condition centroids."""
    labels = np.asarray(labels)
    unique = np.unique(labels)
    centroids = {}
    for g in unique:
        mask = labels == g
        c = u[mask].mean(axis=0)
        c_n = c / max(np.linalg.norm(c), 1e-12)
        centroids[g] = c_n
    pairs = []
    keys = list(centroids.keys())
    for i in range(len(keys)):
        for j in range(i + 1, len(keys)):
            cos = float(centroids[keys[i]] @ centroids[keys[j]])
            pairs.append((keys[i], keys[j], cos))
    return centroids, pairs


def f4_holdout(df, pca_full, labels, param, n_holdout=N_HOLDOUT, n_reps=N_HOLDOUT_REPS, seed=42):
    """P-Khan_F4: stratified holdout, PCA-fit on remaining, nearest-centroid acc."""
    rng = np.random.default_rng(seed)
    n = len(df)
    labels = np.asarray(labels)
    cols = ["Q_max_d0", "R_ohmic_d0", "R_diff_d0",
            "Q_max_d90", "R_ohmic_d90", "R_diff_d90"]
    X_raw = df[cols].values
    accs = []
    for rep in range(n_reps):
        # Stratified holdout
        idx_by_level = {l: np.where(labels == l)[0] for l in np.unique(labels)}
        holdout_idx = []
        for l, idxs in idx_by_level.items():
            n_pick = max(1, int(round(n_holdout * len(idxs) / n)))
            n_pick = min(n_pick, len(idxs) - 1)
            picked = rng.choice(idxs, size=n_pick, replace=False)
            holdout_idx.extend(picked.tolist())
        holdout_idx = np.array(holdout_idx)
        if len(holdout_idx) < 2 or n - len(holdout_idx) < 4:
            continue
        train_idx = np.setdiff1d(np.arange(n), holdout_idx)
        # Fit PCA on train
        X_tr = X_raw[train_idx]
        X_tr_centered = X_tr - X_tr.mean(axis=0, keepdims=True)
        sd_tr = X_tr_centered.std(axis=0, ddof=1)
        sd_tr = np.where(sd_tr < 1e-12, 1e-12, sd_tr)
        z_tr = X_tr_centered / sd_tr
        cov = np.cov(z_tr.T)
        eigvals, eigvecs = np.linalg.eigh(cov)
        comps = eigvecs[:, np.argsort(eigvals)[::-1][:2]]
        pca_tr = z_tr @ comps
        # Train centroids per label
        train_labels = labels[train_idx]
        cents = {}
        for l in np.unique(train_labels):
            mask = train_labels == l
            cents[l] = pca_tr[mask].mean(axis=0)
        # Project test set using train's center + sd + components
        X_te = X_raw[holdout_idx]
        z_te = (X_te - X_tr.mean(axis=0, keepdims=True)) / sd_tr
        pca_te = z_te @ comps
        # Classify test by nearest centroid (Euclidean)
        test_labels = labels[holdout_idx]
        preds = []
        for x in pca_te:
            best = min(cents.keys(), key=lambda l: np.linalg.norm(x - cents[l]))
            preds.append(best)
        acc = float((np.array(preds) == test_labels).mean())
        accs.append(acc)
    return float(np.mean(accs)) if accs else float("nan"), float(np.std(accs)) if accs else float("nan"), len(accs)


def main():
    print(f"=== C3 Khan cross-substrate (pre-reg literature/50, lock a49c830) ===\n")
    df = build_feature_table()
    if len(df) < 10:
        print("ABORT: insufficient cells.")
        return

    pca2, components, explained, z_full, raw_cols = build_pca2_features(df)
    u_full = unit_project(pca2)
    print(f"\nPCA-2 explained variance: {explained:.3f}")
    print(f"PC loadings:")
    for i, c in enumerate(raw_cols):
        print(f"  {c}: PC1={components[i, 0]:+.3f}, PC2={components[i, 1]:+.3f}")

    # Per-design-parameter PERMANOVA
    print(f"\n=== PERMANOVA per design parameter (Bonferroni alpha = {ALPHA_BONFERRONI}) ===")
    results = []
    rng = np.random.default_rng(42)
    for param in ["T_C", "soc_range", "charge_rate"]:
        labels = df[param].astype(str).values
        # Counts per level
        counts = pd.Series(labels).value_counts().to_dict()
        print(f"\n--- {param} ---")
        for k, v in counts.items():
            print(f"  {k}: n={v}")
        # PERMANOVA
        F, p = permanova_test(u_full, labels, rng)
        v = verdict(F, p)
        print(f"  pseudo-F = {F:.3f}")
        print(f"  p (10000 perm) = {p:.4f}")
        print(f"  PERMANOVA VERDICT: {v}")
        # P-Khan_F3 centroid agreement check
        cents, pairs = centroid_check(u_full, labels)
        print(f"  P-Khan_F3 centroid pairwise cosines:")
        for a, b, cos in pairs:
            print(f"    {a} vs {b}: cos = {cos:+.3f}")
        # P-Khan_F4 holdout (only if PERMANOVA PASS or WEAK PASS, per pre-reg)
        if v in ("PASS", "WEAK PASS"):
            f4_acc, f4_sd, f4_n = f4_holdout(df, pca2, labels, param)
            f4_floor_ok = f4_acc > F4_ACC_FLOOR
            print(f"  P-Khan_F4 holdout: mean accuracy = {f4_acc:.3f} +/- {f4_sd:.3f} (n_reps={f4_n}, floor={F4_ACC_FLOOR}) -- {'PASS' if f4_floor_ok else 'FAIL (overfit)'}")
        else:
            f4_acc, f4_sd, f4_n, f4_floor_ok = float("nan"), float("nan"), 0, None
            print(f"  P-Khan_F4: skipped (no PERMANOVA PASS to test)")
        results.append({
            "param": param,
            "F": F,
            "p": p,
            "verdict": v,
            "f4_acc": f4_acc,
            "f4_sd": f4_sd,
            "f4_floor_ok": f4_floor_ok,
            "centroid_pairs": str(pairs),
        })

    summary = pd.DataFrame(results)
    out = OUT_DIR / "khan_cross_substrate_results.parquet"
    summary.to_parquet(out)
    print(f"\nWritten: {out}")

    # Disposition
    print(f"\n========== KHAN CROSS-SUBSTRATE DISPOSITION ==========")
    n_pass = sum(1 for _, r in summary.iterrows() if r["verdict"] == "PASS")
    n_weak = sum(1 for _, r in summary.iterrows() if r["verdict"] == "WEAK PASS")
    f4_results = [r["f4_floor_ok"] for _, r in summary.iterrows() if r["f4_floor_ok"] is not None]
    f4_any_ok = any(r is True for r in f4_results) if f4_results else False

    print(f"  PERMANOVA: {n_pass}/3 PASS + {n_weak}/3 WEAK PASS")
    print(f"  P-Khan_F4 holdout: {'>=1 PASSED param has accuracy > floor' if f4_any_ok else 'all PASSED params overfit or no PASSES'}")

    if n_pass == 0 and n_weak == 0:
        print(f"  ==> KHAN FAILS (P-Khan_F1 fires): amendment does NOT transfer at Khan natural noise.")
        print(f"      Battery substrate should be added to RMD-SRC parent doc as 'framework-resistant' pending diagnosis.")
    elif n_pass >= 2 and f4_any_ok:
        print(f"  ==> KHAN TRANSFERS: amendment is real beyond synthetic.")
        print(f"      Battery substrate should be added to RMD-SRC parent doc as 'Done'.")
    elif n_pass >= 2 and not f4_any_ok:
        print(f"  ==> KHAN OVERFIT (P-Khan_F4 fires): {n_pass} PASS but holdout < floor. Downgrade to PARTIAL.")
    else:
        print(f"  ==> KHAN PARTIAL: {n_pass} PASS, {n_weak} WEAK PASS. Document which parameter(s) survived.")


if __name__ == "__main__":
    main()
